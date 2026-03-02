import argparse
import datetime
import sys

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from tabulate import tabulate
import logging


def log_for_level(self, message, *args, **kwargs):
    if self.isEnabledFor(45):
        self._log(45, message, args, **kwargs)


def log_to_root(message, *args, **kwargs):
    logging.log(45, message, *args, **kwargs)


class Arguments:

    def __init__(self):
        ts = int((datetime.datetime.now()).timestamp())
        self.DATABASE = "data.db"
        self.TYPES_ESSAY = ["MBC", "MIC"]
        self.p = self.get_args()
        logging.addLevelName(45, 'SHOW')
        logging.basicConfig(format='%(asctime)s %(levelname)s :%(message)s',
                            level=logging.DEBUG if self.p.verbose else 45)
        setattr(logging, "SHOW", 45)
        setattr(logging.getLoggerClass(), 'show', log_for_level)
        setattr(logging, 'show', log_to_root)
        self.log = logging.getLogger(__name__)

    def get_args(self):
        parser = argparse.ArgumentParser()
        generic = parser.add_argument_group('Generic arguments')
        imp = parser.add_argument_group('Import and export from / to files')
        db = parser.add_argument_group('Manipulation with database content, set scope for output')

        generic.add_argument("-v", "--verbose", action='store_true', help="Verbose output")
        generic.add_argument("-n", "--no_question", action='store_true', help="Disable approval question")
        generic.add_argument("-D", "--dry_run", action='store_true',
                             help="Dry run: If present -I validate Excel import data. Optional argument is file name with dry run results. use with -v")
        generic.add_argument("-r", "--range", nargs=2,
                             help="For exports and database updates use 2 timestamps as boundaries form ... to. The rounding i.e  2025.12 could be used")
        generic.add_argument("-l", "--list", nargs='+',
                             help="For exports and database updates provide timestamps list .")
        imp.add_argument("-i", "--import_source", type=str,
                         help=f"Import Excel file with data sources to database {self.DATABASE}. To check the content use parameter --dry_run. ")
        imp.add_argument("-s", "--sheets", nargs='+', type=str,
                         help="Source worksheets. If missing all worksheets will be used in  given import file or database")
        imp.add_argument("-I", "--import_backup", type=str,
                         help=f"Import exported data from file to database {self.DATABASE}. To check the content use parameter --dry_run. ")
        imp.add_argument("-E", "--export_backup", type=str,
                         help=f"Import exported data from file to database {self.DATABASE}. To check the content use parameter --dry_run. ")
        imp.add_argument("-e", "--export_final", help=f"Export to final excel file.")
        imp.add_argument("-t", "--type_essay", nargs='+', help="MIC and / or MBC, sheets in final export file ")
        db.add_argument("-d", "--delete", action='store_true', help="Delete records with timestamps by range or list")
        db.add_argument("-j", "--join", action='store_true',
                        help="Join records with timestamps as the list or range (if -r is present). Actual timestamp as is used for joined data")

        return parser.parse_args()

    def check_args(self):
        if len(sys.argv) == 1:
            print("Hi, no expected arguments, let try --help for beginning")
            exit(0)
        if not self.p.type_essay:
            self.p.type_essay = self.TYPES_ESSAY
        if self.p.import_source:
            try:
                wbi = openpyxl.load_workbook(self.p.import_source)
            except  InvalidFileException as e:
                self.log.critical(f"Cannot load file {self.p.import_source}. \n{e}")
                exit(1)
            sheet_ok_manes = []
            self.p.sheets = self.p.sheets if self.p.sheets else wbi.sheetnames
            for sheet_name in self.p.sheets:
                try:
                    rs = wbi[str(sheet_name)].max_row
                    sheet_ok_manes.append(sheet_name)
                    self.log.info(f"Worksheet '{sheet_name}' has {rs} rows.")
                except KeyError:
                    self.log.info(f"Worksheet '{sheet_name}' not exists. It is excluded")
                self.p.sheets = sheet_ok_manes
        self.log.info("List of variables:\n" + tabulate((dict(vars(self.p))).items(), headers=["Variable", "Value"],
                                                        tablefmt="grid"))
        if (not self.p.no_question) and (not input("Do you like to proceed the task? [Y/n]") == "Y"):
            self.log.info("Script terminated by user.")
            exit(0)

    def save_file(self, save_function, file_path: str):
        try:
            save_function(file_path)
        except Exception as e:
            self.log.warning(f"Error saving file to '{file_path}': {e}")
            exit(1)

    def open_file(self, open_function, file_path: str):
        try:
            wbi = open_function(file_path)
        except FileNotFoundError as e:
            self.log.error(e)
            exit(1)
        except PermissionError as e:
            self.log.error(e)
            exit(1)
        return wbi
