import re
from datetime import datetime
from typing import Any

import tabulate
from pandas import DataFrame
import openpyxl
import pandas
from src.bin.arguments import Arguments


def is_empty_integer(v) -> bool:
    return v is None or isinstance(v, int) or v.isdigit()


class ExcelParser(Arguments):
    def __init__(self):
        super().__init__()
        self.ACTIVITIES = ["MIC", "MBC", "MICb", "gentamicin"]
        self.ITEMS = [1, 2, 3, 1, 2, 3, 1, 2, 3, 1]
        self.COLUMNS = ("sheet", "row_id", "code", "pathogen", "activity", "item", "item_value", "timestamp")
        self.COLUMNS_ERR = ("sheet", "cell", "actual value", "error_description")
        self.COLUMNS_INFO = ("sheet", "status" , "row_count", "err_count")
        self.ITEM_COL_OFFSET = 2
        self.CODE_REGEX = r"^\s*\d+\s*-[\s\S]{7,}$"

    @staticmethod
    def get_ts():
        return datetime.now().strftime('%Y-%m-%d_%H:%M:%S')

    def get_db_data(self, wbi: openpyxl.workbook.Workbook) -> pandas.DataFrame:
        timestamp: str = self.get_ts()
        code = ""
        activity = ""
        raw_data = pandas.DataFrame(columns=self.COLUMNS)
        self.log.info(f"There is {len(self.p.sheets)} sheet(s) selected.")
        self.approve_data(wbi)
        for sheet_name in self.p.sheets:
            self.log.info(f"Building data for Excel worksheet {str(sheet_name)}.")
            for row_number in range(1, wbi[str(sheet_name)].max_row):
                lead = wbi[sheet_name].cell(row=row_number, column=2)
                if lead.value:
                    row_id = lead.value
                    if int(lead.value) == 1:
                        raw_code = str(lead.offset(row=-3, column=2).value)
                        code = (raw_code.partition("-")[2]).lstrip().rstrip()
                    pathogen = lead.offset(row=0, column=1).value
                    activity_id = 0
                    for item_id, item in enumerate(self.ITEMS):
                        if item == 1:
                            activity = self.ACTIVITIES[activity_id]
                            activity_id += 1
                        item_v = str(lead.offset(row=0, column=self.ITEM_COL_OFFSET + item_id).value)
                        raw_data.loc[len(raw_data)] = [str(sheet_name), row_id, code, pathogen, activity, item, item_v,timestamp]
        return raw_data

    def approve_data(self, wbi) -> DataFrame:
        data_info = []
        report_err = pandas.DataFrame(columns=self.COLUMNS_ERR)
        bad_sheets = []
        for sheet_name in self.p.sheets:
            if sheet_name in wbi.sheetnames:
                err_count = 0
                row_count = wbi[str(sheet_name)].max_row
                for row_number in range(1, row_count):
                    lead_cell = wbi[sheet_name].cell(row=row_number, column=2)
                    lead = lead_cell.value
                    if not (lead is None or isinstance(lead, int) or lead.isdigit()):
                        report_err.loc[len(report_err)] = [str(sheet_name), f"B{row_number}", str(lead), "Integer number expected"]
                        err_count += 1
                    elif not lead is None and int(lead) == 1:
                        try:
                            raw_code = str(lead_cell.offset(row=-3, column=2).value)
                            if  raw_code is None or not (re.match(self.CODE_REGEX, raw_code)):
                                report_err.loc[len(report_err)] = [str(sheet_name), f"D{row_number - 3}", str(raw_code), "The format of raw code could be '# - code'"]
                                err_count += 1
                        except ValueError as e:
                            report_err.loc[len(report_err)] = [str(sheet_name), f"B{row_number}", str(lead), f"Wrong position of leading '1' {e}"]
                            err_count += 1
                if err_count > 0 :
                    bad_sheets.append(sheet_name)
                    status = "Excluded: data error"
                else:
                    status = "Correct"
            else:
                bad_sheets.append(sheet_name)
                status = "Excluded: not existing name"
                row_count = "Not available"
                err_count = "Not available"
            data_info.append([sheet_name, status, row_count, err_count])
        self.p.sheets = set(self.p.sheets) - set(bad_sheets)
        self.log.info(f"Analysis of {self.p.import_source} file:\n" + tabulate.tabulate(data_info, headers=self.COLUMNS_INFO, tablefmt="grid"))
        return report_err

    def report_errors(self):
        wbi = self.open_file(openpyxl.load_workbook, self.p.import_source)
        self.p.sheets = self.p.sheets if self.p.sheets else wbi.sheetnames
        err_data = self.approve_data(wbi)
        parts = self.p.import_source.rsplit('.', 1)
        err_filename = f"{parts[0]}_errors.{parts[1]}"
        err_data.to_excel(err_filename)
        self.log.info(f"Errors exported to file {err_filename}. Count of errors {len(err_data)}")
